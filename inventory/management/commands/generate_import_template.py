"""Generates the initial-load spreadsheet handed to the shop while the real
bulk-import feature is being built. Pulls its dropdown lists live from the
current catalogs, so re-running this after adding categories/colors/etc.
keeps the template in sync. Not part of the app's runtime: this is an
operator utility, run manually: `manage.py generate_import_template`.
"""

from pathlib import Path

from django.core.management.base import BaseCommand
from django.utils.translation import gettext_lazy as _
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from catalogs.models import ColorVariant, Presentation, ProductSubcategory
from contacts.models import Supplier

DEFAULT_OUTPUT = Path("import_templates/plantilla_inventario_inicial.xlsx")

RUBY = "7A1B2E"
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor=RUBY)
EXAMPLE_FILL = PatternFill("solid", fgColor="EFEFEF")
EXAMPLE_FONT = Font(italic=True, color="666666")
TITLE_FONT = Font(bold=True, size=14, color=RUBY)
LAST_ROW = 500


class Command(BaseCommand):
    help = "Regenerates the initial-inventory-load Excel template from the live catalogs."

    def add_arguments(self, parser):
        parser.add_argument("--output", default=str(DEFAULT_OUTPUT))

    def handle(self, *args, **options):
        category_subcategory_pairs = [
            f"{sub.category.name} / {sub.name}"
            for sub in ProductSubcategory.objects.filter(is_active=True)
            .select_related("category")
            .order_by("category__name", "name")
        ]
        colors = list(
            ColorVariant.objects.filter(is_active=True).order_by("name").values_list("name", flat=True)
        )
        presentations = list(
            Presentation.objects.filter(is_active=True).order_by("name").values_list("name", flat=True)
        )
        suppliers = list(
            Supplier.objects.filter(is_active=True).order_by("name").values_list("name", flat=True)
        )

        workbook = self._build_workbook(category_subcategory_pairs, colors, presentations, suppliers)

        output_path = Path(options["output"])
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        self.stdout.write(self.style.SUCCESS(f"Saved template to {output_path}"))

    def _style_header_row(self, ws, row, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 30

    def _autosize(self, ws, widths):
        for col_letter, width in widths.items():
            ws.column_dimensions[col_letter].width = width

    def _build_workbook(self, category_subcategory_pairs, colors, presentations, suppliers):
        wb = Workbook()

        # ---------------------------------------------------- Instructions
        ws = wb.active
        ws.title = str(_("Instructions"))
        ws["A1"] = f"Joyería Ruby: {_('Initial Inventory Load Template')}"
        ws["A1"].font = TITLE_FONT
        ws.merge_cells("A1:F1")

        instructions = [
            "",
            str(_("How to use this template:")),
            "",
            str(_("1. Fill in the 'Products' sheet, with one row per product (SKU) you want to add.")),
            str(_("   - 'Internal Code' is only for this template: use it to connect each product with")),
            str(_("     its batches in the 'Initial Entries' sheet. It isn't the system's final SKU.")),
            str(_("   - Category/Subcategory, Color, Presentation, and Supplier have a dropdown")),
            str(_("     list. Click the cell and choose an option to avoid mistakes.")),
            str(_("   - If your catalog doesn't have the option you need, add it in the system first")),
            str(_("     and ask for an updated template (or re-run the command that generates it).")),
            "",
            str(_("2. Fill in the 'Initial Entries' sheet, with one row per batch with its own cost.")),
            str(_("   - The same product can have several rows if you bought it at different prices")),
            str(_("     (e.g. 20 units at S/2.00 + 40 units at S/1.50). The system calculates the")),
            str(_("     weighted average cost automatically, so you don't have to calculate it yourself.")),
            str(_("   - Each product's initial stock is simply the sum of its entries here.")),
            str(_("     There's no separate 'initial stock' field.")),
            "",
            str(_("3. The 'Catalogs (reference)' sheet only feeds the dropdown lists. Don't edit it.")),
            "",
            str(_("4. Once the bulk-upload module is ready, this file will be uploadable directly.")),
            str(_("   For now, keep filling it in at your own pace. None of your work is lost.")),
        ]
        for i, line in enumerate(instructions, start=2):
            cell = ws.cell(row=i, column=1, value=line)
            if line[:2].rstrip(".").isdigit():
                cell.font = Font(bold=True)
        self._autosize(ws, {"A": 100})

        # ---------------------------------------------------- Catalogs (reference)
        ref = wb.create_sheet(str(_("Catalogs (reference)")))
        for col, title in enumerate(
            [str(_("Category / Subcategory")), str(_("Color")), str(_("Presentation")), str(_("Supplier"))],
            start=1,
        ):
            ref.cell(row=1, column=col, value=title)
        self._style_header_row(ref, 1, 4)

        for i, value in enumerate(category_subcategory_pairs, start=2):
            ref.cell(row=i, column=1, value=value)
        for i, value in enumerate(colors, start=2):
            ref.cell(row=i, column=2, value=value)
        for i, value in enumerate(presentations, start=2):
            ref.cell(row=i, column=3, value=value)
        for i, value in enumerate(suppliers, start=2):
            ref.cell(row=i, column=4, value=value)
        self._autosize(ref, {"A": 32, "B": 14, "C": 18, "D": 24})

        def defined_name(name, col_letter, count):
            formula = f"'{ref.title}'!${col_letter}$2:${col_letter}${1 + max(count, 1)}"
            wb.defined_names[name] = DefinedName(name, attr_text=formula)

        defined_name("SubcategoryList", "A", len(category_subcategory_pairs))
        defined_name("ColorList", "B", len(colors))
        defined_name("PresentationList", "C", len(presentations))
        defined_name("SupplierList", "D", len(suppliers))

        # ---------------------------------------------------- Products
        prod = wb.create_sheet(str(_("Products")))
        headers = [
            str(_("Internal Code *")),
            str(_("Base Model *")),
            str(_("Category / Subcategory *")),
            str(_("Color")),
            str(_("Presentation")),
            str(_("Supplier")),
            str(_("Suggested Price (S/) *")),
            str(_("Minimum Stock")),
        ]
        for col, title in enumerate(headers, start=1):
            prod.cell(row=1, column=col, value=title)
        self._style_header_row(prod, 1, len(headers))

        example_rows = [
            ["P001", "Aretes Fantasía S/5", "Aretes / Fantasía", "", "Individual", "", 5.00, 10],
            ["P002", "Collar Fina Perla", "Collares / Fina", "Dorado", "Individual", "", 45.00, 3],
        ]
        for r, row in enumerate(example_rows, start=2):
            for c, value in enumerate(row, start=1):
                cell = prod.cell(row=r, column=c, value=value)
                cell.fill = EXAMPLE_FILL
                cell.font = EXAMPLE_FONT

        dv_catsub = DataValidation(type="list", formula1="=SubcategoryList", allow_blank=True)
        dv_color = DataValidation(type="list", formula1="=ColorList", allow_blank=True)
        dv_pres = DataValidation(type="list", formula1="=PresentationList", allow_blank=True)
        dv_supplier = DataValidation(type="list", formula1="=SupplierList", allow_blank=True)
        for dv in (dv_catsub, dv_color, dv_pres, dv_supplier):
            prod.add_data_validation(dv)
        dv_catsub.add(f"C2:C{LAST_ROW}")
        dv_color.add(f"D2:D{LAST_ROW}")
        dv_pres.add(f"E2:E{LAST_ROW}")
        dv_supplier.add(f"F2:F{LAST_ROW}")

        self._autosize(prod, {"A": 16, "B": 26, "C": 26, "D": 14, "E": 16, "F": 20, "G": 20, "H": 14})
        prod.freeze_panes = "A2"

        # ---------------------------------------------------- Initial Entries
        entries = wb.create_sheet(str(_("Initial Entries")))
        for col, title in enumerate(
            [
                str(_("Internal Code *")),
                str(_("Quantity *")),
                str(_("Unit Cost (S/) *")),
                str(_("Notes")),
            ],
            start=1,
        ):
            entries.cell(row=1, column=col, value=title)
        self._style_header_row(entries, 1, 4)

        example_entries = [
            ["P001", 20, 2.00, "Lote 1, proveedor A"],
            ["P001", 40, 1.50, "Lote 2, proveedor B"],
            ["P002", 3, 28.00, "Compra inicial"],
        ]
        for r, row in enumerate(example_entries, start=2):
            for c, value in enumerate(row, start=1):
                cell = entries.cell(row=r, column=c, value=value)
                cell.fill = EXAMPLE_FILL
                cell.font = EXAMPLE_FONT
        self._autosize(entries, {"A": 16, "B": 12, "C": 20, "D": 32})
        entries.freeze_panes = "A2"

        return wb
